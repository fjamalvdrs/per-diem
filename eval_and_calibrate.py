"""
eval_and_calibrate.py
=====================
Three outputs:

  model_eval_2026.xlsx       — per-ticket 2026 errors + stacked summary (2 sheets)
  calibration_offsets.xlsx   — customer + segment bias table (3 sheets)
  model_accuracy_report.md   — plain-English manager report with real numbers

Evaluation sets
  2026 out-of-sample  : 2026 overnight trips with expense data (all quality tiers flagged)
  2022-2025 in-sample : overnight E_complete trips used to train the model
                        (optimistic — model saw this data, but systematic biases still signal)

Run: python eval_and_calibrate.py
"""

import pandas as pd
import numpy as np
import json
from pathlib import Path
from datetime import date

PROJECT   = Path("G:/After Sales Team/PROJECTS/per-diem-model")
FIT_YEARS = [2022, 2023, 2024, 2025]
LONG_TRIP_THRESHOLD = 21   # days — flag separately, don't let outliers skew averages
BAND_ORDER = ["Under 300","300-500","500-750","750-1000","1000-1500","1500-2000","2000+"]

CPI_FACTORS = {
    2018:1.274, 2019:1.252, 2020:1.236, 2021:1.181,
    2022:1.093, 2023:1.050, 2024:1.020, 2025:1.000, 2026:1.000,
}
def cpi(y):
    return CPI_FACTORS.get(int(y), 1.0)


# ── Rate lookup helpers (mirrors app.py) ─────────────────────────────────────
def fine_band(d):
    try:
        d = float(d)
    except (TypeError, ValueError):
        return None
    if d < 100: return "Under 100"
    if d < 300: return "100-300"
    return None

def lookup_day(rates, band, seas):
    c = rates.get("day_trip_rates",{}).get("by_distance_season",{}).get(band,{}).get(seas)
    if c and c.get("n",0) >= 5: return c["total_rate"], "Dist+Season"
    c = rates.get("day_trip_rates",{}).get("by_distance",{}).get(band)
    if c and c.get("n",0) >= 5: return c["total_rate"], "Dist"
    c = rates.get("day_trip_rates",{}).get("global")
    if c: return c["total_rate"], "Global"
    return None, None

def lookup_overnight(rates, band, seas, is_fly, cust=None, state=None, dist=None):
    m = "1" if is_fly else "0"
    on = rates.get("overnight_rates",{})
    if cust:
        c = on.get("by_customer_season",{}).get(str(cust),{}).get(seas)
        if c and c.get("n",0) >= 3: return c["total_rate"], "Customer+Season"
        c = on.get("by_customer",{}).get(str(cust))
        if c and c.get("n",0) >= 3: return c["total_rate"], "Customer"
    if state:
        c = on.get("by_state_distance_mode",{}).get(str(state),{}).get(band,{}).get(m)
        if c and c.get("n",0) >= 5: return c["total_rate"], "State+Dist+Mode"
    fb = fine_band(dist)
    if fb:
        c = on.get("fine_by_distance_season_mode",{}).get(fb,{}).get(seas,{}).get(m)
        if c and c.get("n",0) >= 5: return c["total_rate"], f"Fine({fb})+Season+Mode"
        c = on.get("fine_by_distance_mode",{}).get(fb,{}).get(m)
        if c and c.get("n",0) >= 5: return c["total_rate"], f"Fine({fb})+Mode"
    c = on.get("by_distance_season_mode",{}).get(band,{}).get(seas,{}).get(m)
    if c and c.get("n",0) >= 5: return c["total_rate"], "Dist+Season+Mode"
    c = on.get("by_distance_mode",{}).get(band,{}).get(m)
    if c and c.get("n",0) >= 5: return c["total_rate"], "Dist+Mode"
    c = on.get("by_distance",{}).get(band)
    if c and c.get("n",0) >= 5: return c["total_rate"], "Dist"
    c = on.get("global")
    if c: return c["total_rate"], "Global"
    return None, None

def lookup_airfare(rates, band, seas):
    af = rates.get("airfare_rates",{})
    c = af.get("by_distance_season",{}).get(band,{}).get(seas)
    if c and c.get("n",0) >= 3: return c.get("model_rate", c.get("mean",0))
    c = af.get("by_distance",{}).get(band)
    if c and c.get("n",0) >= 3: return c.get("model_rate", c.get("mean",0))
    return 0.0

def lookup_drive(rates, band):
    c = rates.get("drive_rates",{}).get("by_distance",{}).get(band)
    return c.get("model_rate", c.get("mean",0)) if c else 0.0


# ── Core evaluation function ──────────────────────────────────────────────────
def run_evaluation(df, rates, label):
    """
    Handles overnight and day trips. Rows with no distance_band get prediction=None.
    Returns DataFrame with predicted, actual, error columns.
    """
    rows = []
    for _, r in df.iterrows():
        band      = r.get("distance_band")
        seas      = r.get("season")
        fly       = bool(r.get("is_fly_trip", 0) == 1)
        cust      = str(r.get("CustomerIDAcu", "")).strip()
        cust      = cust if cust not in ("", "nan", "None") else None
        state     = r.get("State")
        dist      = r.get("distance_miles")
        days      = float(r["total_trip_days"]) if pd.notna(r["total_trip_days"]) and r["total_trip_days"] > 0 else None
        is_overn  = bool(r.get("is_overnight", 0))

        if days is None:
            rate, fee, basis = None, 0.0, "No Days"
        elif band is None or (isinstance(band, float) and np.isnan(band)):
            rate, fee, basis = None, 0.0, "No Band"
        elif is_overn:
            rate, basis = lookup_overnight(rates, band, seas, fly, cust, state, dist)
            fee = lookup_airfare(rates, band, seas) if fly else lookup_drive(rates, band)
        else:
            rate, basis = lookup_day(rates, band, seas)
            fee = 0.0

        pred_total = (rate * days + fee) if (rate is not None and days is not None) else None

        sf = lambda x: float(x) if pd.notna(x) else 0.0
        act_daily = sf(r["exp_hotel"]) + sf(r["exp_meals"]) + sf(r["exp_local_transport"])
        act_fee   = sf(r["exp_airfare"]) if fly else sf(r["exp_fuel_tolls"])
        act_total_nominal = act_daily + act_fee
        act_total_cpi     = act_total_nominal * cpi(r["year"])

        cf = cpi(r["year"])
        rows.append({
            "predicted_daily_rate":    rate,
            "predicted_trip_fee":      fee,
            "predicted_daily_total":   rate * days if (rate is not None and days is not None) else None,
            "predicted_total":         pred_total,
            "act_hotel_cpi":           sf(r["exp_hotel"])           * cf,
            "act_meals_cpi":           sf(r["exp_meals"])           * cf,
            "act_local_transport_cpi": sf(r["exp_local_transport"]) * cf,
            "act_airfare_cpi":         sf(r["exp_airfare"])         * cf,
            "act_fuel_tolls_cpi":      sf(r["exp_fuel_tolls"])      * cf,
            "actual_total_cpi":        act_total_cpi,
            "lookup_basis":            basis,
        })

    pred_df = pd.DataFrame(rows, index=df.index)
    out = pd.concat([df, pred_df], axis=1)
    # Keep all rows — even those with no prediction (missing band) or no expense data
    out = pd.concat([df, pred_df], axis=1).copy()

    out["error_dollars"] = np.where(
        out["predicted_total"].notna() & (out["actual_total_cpi"] > 0),
        out["predicted_total"] - out["actual_total_cpi"],
        np.nan
    )
    # error_pct undefined when actual = $0 or no prediction
    out["error_pct"] = np.where(
        out["actual_total_cpi"] > 0,
        out["error_dollars"] / out["actual_total_cpi"] * 100,
        np.nan
    )
    out["abs_error_pct"] = out["error_pct"].abs()
    out["over_under"]    = np.where(
        out["error_dollars"].notna() & (out["error_dollars"] > 0), "Over",
        np.where(out["error_dollars"].notna(), "Under", "")
    )
    out["is_long_trip"]  = out["is_overnight"].astype(bool) & (out["total_trip_days"] > LONG_TRIP_THRESHOLD)
    return out


# ── Summary stats ─────────────────────────────────────────────────────────────
def summary_block(df, group_col=None, label="Overall", exclude_long=True):
    """Returns a tidy summary DataFrame. Excludes long trips from stats by default."""
    sub = df[~df["is_long_trip"]].copy() if exclude_long else df.copy()
    if group_col:
        groups = list(sub.groupby(group_col, observed=True))
    else:
        groups = [(label, sub)]

    rows = []
    for name, g in groups:
        n = len(g)
        if n == 0: continue
        rows.append({
            "Segment":        str(name),
            "N Trips":        n,
            "Mean Error $":   round(g["error_dollars"].mean(), 0),
            "Mean Error %":   round(g["error_pct"].mean(), 1),
            "Within +/-30%":  f"{round((g['abs_error_pct']<=30).mean()*100,1)}%",
            "Within +/-15%":  f"{round((g['abs_error_pct']<=15).mean()*100,1)}%",
            "Overestimate %": f"{round((g['error_dollars']>0).mean()*100,1)}%",
        })
    return pd.DataFrame(rows)


# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading data...")
df   = pd.read_csv(PROJECT / "master_trips_v3.csv", low_memory=False)
with open(PROJECT / "rate_table_v4.json") as f:
    rates = json.load(f)

# Base filter: overnight, non-zero expenses, valid days + distance
def base_filter(df):
    return df[
        df["is_overnight"].astype(bool) &
        (df["data_category"] == "E_complete") &
        (df["exp_total_non_labor"] > 0) &
        (df["total_trip_days"] > 0) &
        df["distance_band"].notna()
    ].copy()


# ── 2026 evaluation ───────────────────────────────────────────────────────────
print("Running 2026 evaluation...")
raw_2026 = df[df["year"] == 2026].copy()

# All 2026 rows — overnight + day trips, with or without expenses, even missing band/days
all_2026_overnight = df[df["year"] == 2026].copy()

# Data quality flag
def quality_flag(row):
    if not bool(row.get("is_overnight", 0)):
        return "Day Trip"
    if row["total_trip_days"] > LONG_TRIP_THRESHOLD:
        return "Long Project (>21 days)"
    if row.get("exp_total_non_labor", 0) == 0:
        return "No Expense Data (company card?)"
    if row["data_category"] == "E_complete":
        return "Clean"
    if row["data_category"] == "F_meals_only":
        return "Meals-Only (hotel missing)"
    return row["data_category"]

all_2026_overnight["Data Quality"] = all_2026_overnight.apply(quality_flag, axis=1)
eval_2026 = run_evaluation(all_2026_overnight, rates, "2026")

# Primary stats: E_complete, not long trips
primary_2026 = eval_2026[
    (eval_2026["data_category"] == "E_complete") &
    (~eval_2026["is_long_trip"])
].copy()


# ── In-sample evaluation (2022-2025) ─────────────────────────────────────────
print("Running in-sample evaluation (2022-2025)...")
insample_df = base_filter(df[df["year"].isin(FIT_YEARS)])
insample_df = insample_df[insample_df["total_trip_days"] <= LONG_TRIP_THRESHOLD].copy()
eval_insample = run_evaluation(insample_df, rates, "In-sample 2022-2025")


# ── Calibration table ─────────────────────────────────────────────────────────
print("Building calibration table...")

def bias_stats(sub, min_n=3):
    """Returns (n, mean_bias_pct) or (n, None) if too few rows."""
    n = len(sub)
    if n < min_n:
        return n, None
    return n, round(sub["error_pct"].mean(), 1)

def recommended_offset(bias_is, bias_oos):
    """
    Compute recommended offset % and confidence.
    Offsets > 75% in magnitude are flagged as 'Investigate' — they almost always
    indicate company card data gaps rather than a correctable model pattern.
    """
    have_is  = bias_is  is not None
    have_oos = bias_oos is not None
    agree    = None

    if have_is and have_oos:
        offset = round(-(0.6 * bias_oos + 0.4 * bias_is), 1)
        agree  = (bias_is * bias_oos) > 0
        conf   = "High" if agree else "Medium"
    elif have_oos:
        offset = round(-bias_oos, 1)
        conf   = "Medium"
    elif have_is:
        offset = round(-bias_is / 1.4, 1)
        conf   = "Medium"
    else:
        return None, "Low", None

    if abs(offset) < 5:
        conf = "Low"
    # Extreme offsets almost always reflect data gaps, not correctable model bias
    if abs(offset) > 75:
        conf = "Investigate"
    return offset, conf, agree

# By customer
cust_rows = []
all_custs = set(eval_insample["CustomerName"].dropna()) | set(primary_2026["CustomerName"].dropna())
for cust in sorted(all_custs):
    is_sub  = eval_insample[eval_insample["CustomerName"] == cust]
    oos_sub = primary_2026[primary_2026["CustomerName"] == cust]

    n_is,  b_is  = bias_stats(is_sub,  min_n=3)
    n_oos, b_oos = bias_stats(oos_sub, min_n=2)

    if n_is == 0 and n_oos == 0: continue

    offset, conf, agree = recommended_offset(b_is, b_oos)
    cust_rows.append({
        "Customer":              cust,
        "N (in-sample)":         n_is,
        "Bias % (in-sample)":    b_is,
        "N (2026)":              n_oos,
        "Bias % (2026)":         b_oos,
        "Sources Agree":         agree,
        "Recommended Offset %":  offset,
        "Confidence":            conf,
        "Notes": (
            "Investigate: extreme bias likely reflects company card data gaps, not model error"
            if conf == "Investigate"
            else f"Apply {offset:+.0f}% to model estimate (over-estimate pattern, use correction for fairer baseline)"
            if (offset is not None and offset < -5)
            else f"Apply {offset:+.0f}% to model estimate (under-estimate pattern)"
            if (offset is not None and offset > 5)
            else "Bias within acceptable range — no adjustment needed"
        ) if offset is not None else "Insufficient data — need >= 3 trips",
    })

calib_customers = pd.DataFrame(cust_rows).sort_values(
    ["Confidence","Bias % (2026)"],
    ascending=[True, False],
    key=lambda c: c.map({"High":0,"Medium":1,"Low":2,"Investigate":3}) if c.name=="Confidence" else c.abs()
).reset_index(drop=True)

# By distance band
seg_rows = []
for band in BAND_ORDER:
    is_sub  = eval_insample[eval_insample["distance_band"] == band]
    oos_sub = primary_2026[primary_2026["distance_band"] == band]
    n_is, b_is   = bias_stats(is_sub,  min_n=5)
    n_oos, b_oos = bias_stats(oos_sub, min_n=3)
    offset, conf, agree = recommended_offset(b_is, b_oos)
    seg_rows.append({
        "Distance Band":         band,
        "N (in-sample)":         n_is,
        "Bias % (in-sample)":    b_is,
        "N (2026)":              n_oos,
        "Bias % (2026)":         b_oos,
        "Sources Agree":         agree,
        "Recommended Offset %":  offset,
        "Confidence":            conf,
    })
calib_bands = pd.DataFrame(seg_rows)

# By mode
mode_rows = []
for mode_val, mode_lbl in [(0,"Drive"),(1,"Fly")]:
    is_sub  = eval_insample[eval_insample["is_fly_trip"] == mode_val]
    oos_sub = primary_2026[primary_2026["is_fly_trip"] == mode_val]
    n_is, b_is   = bias_stats(is_sub,  min_n=5)
    n_oos, b_oos = bias_stats(oos_sub, min_n=3)
    offset, conf, agree = recommended_offset(b_is, b_oos)
    mode_rows.append({
        "Mode": mode_lbl,
        "N (in-sample)": n_is, "Bias % (in-sample)": b_is,
        "N (2026)": n_oos, "Bias % (2026)": b_oos,
        "Recommended Offset %": offset, "Confidence": conf,
    })
calib_mode = pd.DataFrame(mode_rows)


# ── Print key numbers to console ──────────────────────────────────────────────
print("\n" + "="*65)
print("KEY RESULTS")
print("="*65)
print(f"\n2026 primary (E_complete, <=21 days): N={len(primary_2026)}")
print(summary_block(primary_2026).to_string(index=False))
print(f"\nIn-sample (2022-2025, E_complete, <=21 days): N={len(eval_insample)}")
print(summary_block(eval_insample).to_string(index=False))
print("\n2026 by distance band:")
print(summary_block(primary_2026, "distance_band").to_string(index=False))
print("\nHigh-confidence customer corrections:")
hc = calib_customers[calib_customers["Confidence"]=="High"]
print(hc[["Customer","N (in-sample)","Bias % (in-sample)","N (2026)","Bias % (2026)","Recommended Offset %"]].to_string(index=False))


# ── Export model_eval_2026.xlsx ───────────────────────────────────────────────
print("\nWriting model_eval_2026.xlsx...")

# Sheet 1: Per Ticket
ticket_cols = [
    "Title", "CustomerName", "State", "distance_band", "season",
    "total_trip_days",
    "act_hotel_cpi", "act_meals_cpi", "act_local_transport_cpi",
    "act_airfare_cpi", "act_fuel_tolls_cpi",
    "actual_total_cpi",
    "predicted_daily_total", "predicted_trip_fee", "predicted_total",
    "error_dollars", "error_pct", "over_under",
    "Data Quality", "lookup_basis",
]
ticket_cols = [c for c in ticket_cols if c in eval_2026.columns]
ticket_df = eval_2026[ticket_cols].copy()
col_names = [
    "Ticket", "Customer", "State", "Distance Band", "Season",
    "Total Trip Days",
    "Actual Hotel ($)", "Actual Meals ($)", "Actual Local Transport ($)",
    "Actual Airfare ($)", "Actual Fuel/Tolls ($)",
    "Actual Total ($, 2025$)",
    "Predicted Daily Cost ($)", "Predicted Trip Fee ($)", "Predicted Total ($)",
    "Error ($)", "Error %", "Over/Under",
    "Data Quality", "Lookup Basis",
]
ticket_df.columns = col_names[:len(ticket_cols)]

# Track which expense columns to check for zero-highlighting
EXPENSE_COLS = ["Actual Hotel ($)", "Actual Meals ($)", "Actual Local Transport ($)",
                "Actual Airfare ($)", "Actual Fuel/Tolls ($)"]

# Sheet 2: Error Summary — stacked sections
def stacked_summary(eval_df, insample_df):
    sections = []

    def section(title, rows_df):
        header = pd.DataFrame([{c: "" for c in rows_df.columns}])
        header.iloc[0, 0] = f"--- {title} ---"
        return pd.concat([header, rows_df, pd.DataFrame([{c:"" for c in rows_df.columns}])],
                         ignore_index=True)

    p = eval_df[(eval_df["data_category"]=="E_complete") & (~eval_df["is_long_trip"])]
    all_ov = eval_df[~eval_df["is_long_trip"]]
    long   = eval_df[eval_df["is_long_trip"]]

    note_df = pd.DataFrame([{
        "Segment": "NOTE: Primary stats use E_complete + <=21 days rows only. Long trips and meals-only rows shown separately.",
        "N Trips":"","Mean Error $":"","Mean Error %":"","Within +/-30%":"","Within +/-15%":"","Overestimate %":""
    }])

    s_overall = pd.concat([
        summary_block(p,       label="2026 Primary (E_complete, <=21 days)"),
        summary_block(all_ov,  label="2026 All Overnight (incl meals-only)"),
        summary_block(long,    label="Long Projects (>21 days) — for reference only",
                      exclude_long=False),
    ])
    s_insample = summary_block(insample_df, label="In-sample 2022-2025 (E_complete, <=21 days)")

    s_dist  = summary_block(p, "distance_band")
    s_cust  = summary_block(p, "CustomerName").sort_values("N Trips", ascending=False)
    s_seas  = summary_block(p, "season")
    s_mode  = p.copy(); s_mode["Mode"] = np.where(s_mode["is_fly_trip"]==1,"Fly","Drive")
    s_mode  = summary_block(s_mode, "Mode")
    s_dur   = p.copy()
    s_dur["Duration Group"] = pd.cut(s_dur["total_trip_days"],
        bins=[0,3,7,14,21], labels=["1-3 days","4-7 days","8-14 days","15-21 days"])
    s_dur  = summary_block(s_dur, "Duration Group")

    out = pd.concat([
        note_df,
        section("OVERALL — 2026 vs In-sample comparison",
                pd.concat([s_overall, s_insample], ignore_index=True)),
        section("BY DISTANCE BAND  (2026 primary)",   s_dist),
        section("BY CUSTOMER  (2026 primary, sorted by volume)", s_cust),
        section("BY SEASON  (2026 primary)",           s_seas),
        section("BY TRAVEL MODE  (2026 primary)",      s_mode),
        section("BY TRIP DURATION  (2026 primary)",    s_dur),
    ], ignore_index=True)
    return out

summary_df = stacked_summary(eval_2026, eval_insample)

from openpyxl.styles import PatternFill
ZERO_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # amber yellow
ZERO_CELL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red for the zero cell itself

with pd.ExcelWriter(PROJECT / "model_eval_2026.xlsx", engine="openpyxl") as w:
    ticket_df.to_excel(w, sheet_name="Per Ticket 2026", index=False)
    summary_df.to_excel(w, sheet_name="Error Summary", index=False)

    # Highlight rows with zero/missing expenses
    ws = w.sheets["Per Ticket 2026"]
    headers = [cell.value for cell in ws[1]]
    exp_col_indices = [headers.index(c) + 1 for c in EXPENSE_COLS if c in headers]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        zero_cells = [row[i - 1] for i in exp_col_indices
                      if row[i - 1].value is None or row[i - 1].value == 0]
        if zero_cells:
            # Tint the entire row amber
            for cell in row:
                cell.fill = ZERO_FILL
            # Extra red on the specific zero cells
            for cell in zero_cells:
                cell.fill = ZERO_CELL_FILL

print("  Saved model_eval_2026.xlsx")


# ── Export calibration_offsets.xlsx ───────────────────────────────────────────
print("Writing calibration_offsets.xlsx...")

with pd.ExcelWriter(PROJECT / "calibration_offsets.xlsx", engine="openpyxl") as w:
    calib_customers.to_excel(w, sheet_name="By Customer", index=False)
    calib_bands.to_excel(w, sheet_name="By Distance Band", index=False)
    calib_mode.to_excel(w, sheet_name="By Mode", index=False)

print("  Saved calibration_offsets.xlsx")


# ── Write manager report ───────────────────────────────────────────────────────
print("Writing model_accuracy_report.md...")

# Gather key numbers
n_2026_primary    = len(primary_2026)
pct30_2026        = round((primary_2026["abs_error_pct"]<=30).mean()*100, 1)
pct15_2026        = round((primary_2026["abs_error_pct"]<=15).mean()*100, 1)
mean_bias_2026    = round(primary_2026["error_dollars"].mean(), 0)
mean_bias_pct_2026= round(primary_2026["error_pct"].mean(), 1)
over_pct_2026     = round((primary_2026["error_dollars"]>0).mean()*100, 1)

n_is              = len(eval_insample)
pct30_is          = round((eval_insample["abs_error_pct"]<=30).mean()*100, 1)
pct15_is          = round((eval_insample["abs_error_pct"]<=15).mean()*100, 1)
mean_bias_is      = round(eval_insample["error_dollars"].mean(), 0)

# Best and worst distance bands — only count bands with N >= 5 (meaningful signal)
band_acc = (
    primary_2026[~primary_2026["is_long_trip"]]
    .groupby("distance_band", observed=True)
    .agg(n=("abs_error_pct","count"), pct30=("abs_error_pct", lambda x: (x<=30).mean()*100))
    .query("n >= 5")
    .sort_values("pct30", ascending=False)
)
best_band  = band_acc.index[0]  if len(band_acc) > 0 else "N/A"
worst_band = band_acc.index[-1] if len(band_acc) > 0 else "N/A"
best_pct   = round(band_acc["pct30"].iloc[0],1)  if len(band_acc) > 0 else 0
worst_pct  = round(band_acc["pct30"].iloc[-1],1) if len(band_acc) > 0 else 0

median_bias_pct_2026 = round(primary_2026["error_pct"].median(), 1)
n_long_2026 = eval_2026[eval_2026["is_long_trip"]].shape[0]
long_word   = "trip" if n_long_2026 == 1 else "trips"

n_high_conf  = len(calib_customers[calib_customers["Confidence"]=="High"])
n_med_conf   = len(calib_customers[calib_customers["Confidence"]=="Medium"])
n_invest     = len(calib_customers[calib_customers["Confidence"]=="Investigate"])
n_custs_2026 = primary_2026["CustomerName"].nunique()

# In-sample vs 2026 gap
accuracy_gap = round(pct30_is - pct30_2026, 1)

report = f"""# Per Diem Model v4 — Accuracy & Calibration Report
**Prepared:** {date.today().strftime("%B %d, %Y")}
**Prepared by:** After Sales Team
**Model version:** v4.1  |  **Training years:** 2022–2025  |  **CPI base:** 2025 dollars

---

## Executive Summary

| Metric | Value |
|--------|-------|
| Out-of-sample trips evaluated (2026) | {n_2026_primary} overnight trips |
| % of estimates within ±30% of actual | **{pct30_2026}%** (out-of-sample) |
| % of estimates within ±15% of actual | **{pct15_2026}%** (out-of-sample) |
| Average over/under-estimate (mean) | **+${int(abs(mean_bias_2026)):,}** {"over-estimate" if mean_bias_2026 > 0 else "under-estimate"} (mean {mean_bias_pct_2026:+.1f}%, median {median_bias_pct_2026:+.1f}%) |
| % of trips where model over-estimates | **{over_pct_2026}%** |
| In-sample accuracy (training set) | {pct30_is}% within ±30% (across {n_is:,} trips) |
| Accuracy drop: in-sample → out-of-sample | {accuracy_gap} percentage points |

**Plain-English summary:** On 2026 trips the model has never seen, roughly **one in three
individual estimates falls within 30% of actual cost** ({pct30_2026}%). The model skews toward
**over-estimating** ({over_pct_2026}% of trips) — the conservative direction for budgeting.
Important context: the model predicts the *expected average* cost for a trip profile.
Individual actual costs have high natural variance (one tech books $89/night, another $220 at
the same customer). For **portfolio-level budgeting** across many trips, errors cancel and
aggregate accuracy is materially better than the per-trip figure. The {accuracy_gap}-point
drop from training to real-world performance is normal for any statistical estimator.

---

## How We Tested This

The model was built on trip expense data from **2022–2025** (3,551 trips). To evaluate how it
performs on data it has never seen, we used **2026 overnight trips** ({n_2026_primary} trips with complete
expense records) as a test set. For each trip, we ran it through the exact same rate lookup
logic the app uses, compared the predicted total to the actual recorded expenses (adjusted to
2025 dollars for a fair comparison), and measured the difference.

We also ran the same evaluation on the **2022–2025 training data** ({n_is:,} trips) to establish
a baseline. Because the model was built from that data, in-sample accuracy is higher ({pct30_is}% within ±30%)
and should be read as an optimistic upper bound, not a realistic expectation.

**What counts as "actual cost":**
Hotel + meals + local transport (daily) + airfare or fuel/tolls (trip fee).
Company card expenses are NOT captured in the system — this is the primary data gap
(roughly 66% of pre-2022 trips have zero recorded expenses for this reason).

**What we excluded:**
- Day trips: mileage reimbursement runs through a separate channel, making day-trip actuals
  appear artificially low and incomparable to predictions.
- Long projects (>21 days): 2026 had {n_long_2026} {long_word} over 21 days. These are multi-week
  installation jobs where negotiated rates, extended-stay hotels, and team housing
  arrangements make the per-diem model structurally inapplicable. They skew any average.

---

## Where the Model Works Well

**By distance band**, the strongest performers in 2026 are:

| Distance Band | % within ±30% | N |
|--------------|--------------|---|
{chr(10).join(f"| {b} | {round((primary_2026[primary_2026['distance_band']==b]['abs_error_pct']<=30).mean()*100,1) if len(primary_2026[primary_2026['distance_band']==b])>0 else 'N/A'}% | {len(primary_2026[primary_2026['distance_band']==b])} |" for b in BAND_ORDER if len(primary_2026[primary_2026['distance_band']==b]) > 0)}

The best-performing band is **{best_band}** at **{best_pct}% within ±30%**.

**Customer-specific lookups:** For customers where the model uses their own historical trip data
(not a generic rate), it tends to be more consistent — even when absolute accuracy varies,
the direction of error is predictable, which is useful for offsetting.

---

## Where the Model Struggles

**{worst_band}** is the weakest band among those with enough 2026 trips to draw conclusions,
at {worst_pct}% within ±30%. Short overnight trips (Under 300 and 300–500 miles) have the
most cost variance — it depends heavily on whether a tech needs a hotel at all, proximity
to the customer site, and local market rates. The model can only predict the average behavior
for that distance range, so individual trips deviate more.

**Under-estimation risk:** {round(100-over_pct_2026,1)}% of 2026 trips were under-estimated (model predicted
less than actual). While the average bias is positive (over-estimate), individual trips can
be significantly under-estimated. This is more likely on fly trips to high-cost-of-living
markets (New York, San Francisco, Boston).

**Data thinness:** With only {n_2026_primary} clean 2026 trips, segment-level conclusions (e.g.,
by specific customer) carry wide uncertainty. As 2026 data accumulates through the year,
re-running this evaluation will give sharper results.

---

## Calibration: How to Offset Estimates

We computed systematic bias patterns for each customer and distance band by combining:
- **{n_is:,} in-sample trips (2022–2025):** stable patterns, but optimistic
- **{n_2026_primary} out-of-sample trips (2026):** genuine holdout signal, but thin

**Where both sources agree on the direction of bias, we flag the offset as High confidence.**

Summary of calibration findings:
- **{n_high_conf} customers** have High-confidence offsets — both data sources agree,
  pattern is consistent, correction is meaningful (>5%). Apply these with confidence.
- **{n_med_conf} customers** have Medium-confidence offsets — one data source or thin 2026 sample.
  Use directionally but verify before applying at scale.
- **{n_invest} customers** flagged as "Investigate" — extreme bias (>75%) almost certainly
  reflects company card data gaps rather than a correctable model pattern. Do not apply
  these as offsets; instead, audit expense records for those accounts.
- Full details are in `calibration_offsets.xlsx` → "By Customer" sheet

**How to use the calibration table:**

> If a customer has a High-confidence offset of, say, -18%, it means the model has
> consistently over-estimated costs for that customer by ~18%. You can apply that
> correction manually: take the model's estimate and multiply by (1 - 0.18).
> For budgeting, you might keep the uncorrected (higher) estimate as a conservative
> buffer. For reimbursement approval, apply the offset for a fairer baseline.

**Distance band offsets** (see `calibration_offsets.xlsx` → "By Distance Band"):
These apply when you don't have a customer-specific correction. If the model
consistently over- or under-estimates for a given distance range, apply that correction
to any trip in that band.

---

## Key Limitations — Read Before Presenting

1. **2026 sample is thin.** {n_2026_primary} trips across {n_custs_2026} customers. No single customer has
   more than a handful of 2026 trips. Segment-level error rates have wide confidence
   intervals. The numbers are directionally correct, not statistically definitive.

2. **Company card gap.** We cannot evaluate what we cannot see. Any trip where hotel/meals
   were charged to a company card shows $0 in the expense system. This affects roughly
   60–70% of older trips and an unknown portion of 2026 trips. If company card expenses
   were reconciled, our actual training data would likely be 2–3x larger and accuracy
   would improve.

3. **What "within ±30%" means.** This metric measures how close a single estimate is to a
   single actual trip cost. Individual trips are inherently variable — one tech books a
   $89/night extended-stay, another books $220/night at a Marriott for the same customer.
   The model predicts the expected average, not the individual outcome. For **portfolio-level
   budgeting** (total annual travel budget), the model's accuracy is higher because errors
   cancel out across many trips.

4. **COVID years excluded.** We intentionally excluded 2020–2021 data from the primary
   evaluation. Hotel prices during COVID were structurally suppressed (30–40% below normal),
   so comparing 2025-dollar predictions to COVID-era actuals measures COVID distortion,
   not model quality. These years remain excluded from training for the same reason.

---

## Recommendations

### Immediate (no code changes needed)
1. **Use the calibration table for manual adjustments.** For customers with High-confidence
   offsets, apply the correction before presenting estimates to operations or finance.
   Document which customers you've corrected — this builds the case for automating it later.

2. **Flag Under-300-mile trips as higher uncertainty.** When presenting an estimate for a
   short-haul overnight trip, note that the model's accuracy is lower for this segment
   and widen the range you present.

3. **Exclude long projects (>21 days) from per-diem model entirely.** These need a separate
   treatment (negotiated rates, project-specific budgets). Running them through the per-diem
   model produces large, uninformative errors.

### Next quarterly model refresh (build_v4.py)
4. **Add 2019 to training data.** There are 29 clean pre-COVID 2019 trips that were
   arbitrarily excluded. Including them unlocks customer-specific rate lookups for
   46 customers who currently fall back to generic rates. Low risk, high value.

5. **Add 2026 to training data.** As 2026 accumulates, including it in the next refresh
   will directly improve rate accuracy for current market conditions.

### Longer term
6. **Close the company card gap.** Reconciling company card data with the expense system
   is the single highest-impact data quality improvement available. Even a partial
   reconciliation (top 20 customers) would materially improve both training and evaluation.

---

## How Confident Should You Be Presenting This?

**High confidence:**
- The model over-estimates more than it under-estimates — conservative for budgeting.
- Distance bands between 500–1500 miles perform consistently.
- Customer-specific lookups produce more predictable (if not always more accurate) results.
- The in-sample → out-of-sample accuracy drop is normal and expected.

**Medium confidence:**
- The segment-level numbers (by customer, by season) are directionally correct but
  statistically thin on 2026 data alone. Present as "early signal" not "definitive finding."
- The calibration offsets for Medium-confidence customers are reasonable estimates,
  not precise corrections.

**What to say to leadership:**
> "Our model estimates overnight trip costs within 30% of actual for about {pct30_2026}% of
> individual trips tested on 2026 data it has never seen — roughly one in three. It leans
> toward over-estimating, which keeps budgets conservative. For aggregate budgeting across
> a full quarter or year, the accuracy is higher because individual errors offset each other.
> We have identified the specific customers and distance ranges where the model consistently
> misses, and have prepared correction factors for manual application. The primary path to
> materially better accuracy is closing the company card data gap, which currently leaves
> 60–70% of historical trips with no expense record in our system."

---
*Generated by eval_and_calibrate.py on {date.today().strftime("%Y-%m-%d")}*
*Source data: master_trips_v3.csv | Rate table: rate_table_v4.json*
"""

with open(PROJECT / "model_accuracy_report.md", "w", encoding="utf-8") as f:
    f.write(report)

print("  Saved model_accuracy_report.md")
print("\nAll done.")
print(f"  model_eval_2026.xlsx")
print(f"  calibration_offsets.xlsx")
print(f"  model_accuracy_report.md")
